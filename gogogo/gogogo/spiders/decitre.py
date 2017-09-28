import scrapy
import glob
import os
from gogogo.items import GogogoItem
from openpyxl import load_workbook ,Workbook

import json

URLDESTIMAGELD = '.'
URLDESTPHOTOAD = ''
RECHERCHE = './/*[@id=\'search\']'
FICHETECHD = './/*[@id=\'fiche-technique\']/div[2]/ul/descendant::*/text()'
TITRED = './/div[3]/div[2]/div[2]/h1/text()'
PRIX = './/*[@itemprop=\'price\']/text()'
AUTEURD = './/*[@class=\'authors\']/a/text()'
PREFACED = './/*[@id=\'resume\']/div[2]/text()'
BIOGR = './/*[@id=\'auteur\']/div[2]'
IMAGEL = './/div[1]/div[3]/div[2]/div[1]/div[1]/a/img/@src'
IMAGEA = './/*[@id=\'auteur\']/div[2]/img/@src'
THEMES = './/div[3]/div[5]/div/ul/descendant::*/text()'

fiJdid = '/path/to/file.xlsx'
xlsx_workbook = load_workbook(filename = fiJdid, read_only = True,   use_iterators = True).active
row = 1
mapping  = []

for row in range(1, xlsx_workbook.max_row):
    mapping.append({'url':'http://www.decitre.fr/livres/'+str(xlsx_workbook.cell(row = row, column = 1).value)+'.html','qte':xlsx_workbook.cell(row = row, column = 2).value,'izza_code':xlsx_workbook.cell(row = row, column = 3).value,'ean':xlsx_workbook.cell(row = row, column = 1).value})
def nettoyer(elm):
    return (elm.replace('\n','').strip())
class DecitreSpider(scrapy.Spider):
    name = "decitre"
    allowed_domains = ['www.decitre.fr']
    mapping=mapping
    def lireConf(self):
        with open('/path/to/config.json', 'r') as f:
            config = json.load(f)
            return config
			
    def lireExcel(self):
        pass
    def start_requests(self):
	    headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:48.0) Gecko/20100101 Firefox/48.0'}
        for elm in self.mapping:
            yield scrapy.Request(elm['url'],errback=self.parse_err, callback=self.parse, meta={'izza_code': elm['izza_code'],'qte':elm['qte'],'ean':elm['ean']},headers=headers)
		
    def parse_err(self, response):
        livre = GogogoItem()
        livre['quantite'] = response.meta['qte']
        livre['ean'] = response.meta['ean']
        yield livre
    def parse(self, response):
        livre = GogogoItem()
        livre['url'] = response.url
        livre['quantite'] = response.meta['qte']
        livre['ean'] = response.meta['ean']
        livre['source'] = 'decitre'
        livre['titre'] = response.xpath(TITRED).extract()
        for i in range(len(livre['titre'])):
            livre['titre'][i] = nettoyer(livre['titre'][i])
        livre['prix'] = response.xpath(PRIX).extract()
        livre['preface'] = response.xpath(PREFACED).extract()
        livre['auteur'] = response.xpath(AUTEURD).extract()
        livre['auteur'] = set(livre['auteur'])
        livre['auteur'] = list(livre['auteur'])
        livre['titre'] = set (livre['titre'])
        livre['titre'] = list(livre['titre'])
        for elm in livre['titre']:
            if elm == '':
                livre['titre'].remove(elm)
        fichTechs = response.xpath(FICHETECHD).extract()
        for i in range(len(fichTechs)):
            fichTechs[i] = fichTechs[i].replace('\n','').strip()

        fichTech = filter(None,fichTechs)
        print(fichTech)
        i = iter(fichTech)
        dicos = dict(zip(i,i))
        for k,v in dicos.items():
            if k == 'Format :':
                livre['format_livre'] = nettoyer(v)
            elif k == 'Nb. de pages :':
                livre['nombre_pages'] = nettoyer(v)
            elif k == 'Date de parution :':
                livre['date_parution'] = nettoyer(v)
            elif k == 'ISBN :':
                livre['isbn'] = nettoyer(v)
            elif k == 'Editeur :':
                livre['edition'] = nettoyer(v)
            elif k == 'Présentation :':
                livre['presentation'] = nettoyer(v)
            elif k == 'Poids :':
                livre['poids'] = nettoyer(v)
            elif k == 'Dimensions :':
                livre['dimension'] = nettoyer(v)
            elif k == 'Collection :':
                livre['collection'] = nettoyer(v)

        themes = response.xpath(THEMES).extract()
        for i in range(len(themes)):
            themes[i] = themes[i].replace('\n','')
            themes[i] = themes[i].replace('>','')
            themes[i] = themes[i].strip()
        for i in range(len(themes)):
            themes[i]=nettoyer(themes[i])
        themes=set(themes)
        livre['themes'] = list(themes)
        for elm in livre['themes']:
            if elm == '' or elm == '>':
                livre['themes'].remove(elm)

        livre['image_thumb'] = response.xpath(IMAGEL).extract()
        if(len(livre['image_thumb']) > 0):
            livre['url_image_thumb'] = livre['image_thumb'][0]
            livre['url_image_full'] = livre['url_image_thumb'].replace('200x303','475x500')
        yield livre
